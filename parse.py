import time
import re
import math

from openpyxl import load_workbook
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

from bs4 import BeautifulSoup

from settings import filter_product, control_discount, filter_phone


class Parse:
    def __init__(self, url, category, driver):
        self.price_dict = {}
        self.category = category
        self.url = url
        self.brands_pl = ['Apple', 'Huawei', 'Realme', 'Samsung', 'Xiaomi', 'Honor']
        self.brands_kof = ['DeLonghi', 'Philips', 'Smeg', 'Polaris', 'Nivona', 'Jura', 'Bosh']
        self.driver = driver

    def _get_url(self):
        print(self.url)
        self.driver.set_page_load_timeout(600)
        self.driver.get(self.url)
        self.driver.implicitly_wait(10)
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

    def parse(self):
        self._get_url()
        page = self.driver.page_source
        soup = BeautifulSoup(page, 'lxml')
        pr = soup.find_all('div', {'class': 'item-block'})
        for i in pr:
            product = i.find('div', {'class': 'item-title'}).a
            url = f"https://megamarket.ru{product['href']}"
            name = product['title']
            p = i.find('div', {'class': 'item-money'})
            price = int(re.sub(r"[\s,₽]", "", p.find('span').get_text()))
            try:
                shop = i.find('span', {'class': 'merchant-info__name'}).get_text().strip()
            except AttributeError:
                shop = '0'
            try:
                bonus = int(re.sub(r"[\s,₽]", "", p.find('span', {'class': 'bonus-amount'}).get_text()))
            except AttributeError:
                bonus = 0
            sale = math.ceil((bonus * 100) / price)
            # if sale >= 44 and any(brand in name for brand in self.brands):
            if sale >= control_discount and price > 3000:
                if self.category in ['planshety', 'kofemashiny', 'umnye-chasy', 'smartfony-android']:
                    if self.category == 'smartfony-android':
                        brand = filter_phone(name)
                        if brand:
                            if brand in self.price_dict:
                                self.price_dict[f'{brand}'] |= {name: [url, sale, price, shop]}
                            else:
                                self.price_dict[f'{brand}'] = {name: [url, sale, price, shop]}
                    elif filter_product(name, self.category):
                        self.price_dict[f'{name}'] = [url, sale, price, shop]
                else:
                    self.price_dict[f'{name}'] = [url, sale, price, shop]
                    print('ok')
        return self.price_dict


class Parse_Price:
    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.products = {}
        self.counter = 0
        global wb
        wb = load_workbook(f'./price.xlsx')

    def _set_up(self):  # запускаем браузер
        options = Options()
        user_agent = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
        options.add_argument(f'user-agent={user_agent}')
        options.add_argument('--headless')
        options.add_argument('--start-maximized')
        self.driver = webdriver.Chrome(options=options)

    def _get_url(self, url):
        self.driver.get(f'{url}#?details_block=prices')
        self.driver.implicitly_wait(10)
        self.driver.set_page_load_timeout(600)
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))

    def open(self):
        sheet = wb[f'{self.sheet_name}']
        row = self.counter
        self.products = {}
        for i in range(30):
            row += 1
            if sheet.cell(row=row, column=1).value is None:
                break
            self.products[sheet.cell(row=row, column=1).value] = [sheet.cell(row=row, column=2).value, sheet.cell(row=row, column=3).value]

    def save(self):
        column = self.counter
        sheet = wb[f'{self.sheet_name}_1']
        for name, url_name in self.products.items():
            column += 1
            sheet[f'A{column}'], sheet[f'B{column}'], sheet[f'C{column}'] = name, url_name[0], url_name[1]
        wb.save('price.xlsx')

    def create_table(self):
        print('create')
        if f'{self.sheet_name}_1' in wb.sheetnames:
            wb.remove(wb[f'{self.sheet_name}_1'])
        wb.create_sheet(f'{self.sheet_name}_1')
        wb.save('./price.xlsx')

    def calculation(self):
        self.open()
        for name, url in self.products.copy().items():
            self._set_up()
            self._get_url(url[0])
            page = self.driver.page_source
            soup = BeautifulSoup(page, 'lxml')
            text = soup.find_all('div', {'class': 'product-offer product-offer_with-payment-method'})
            price_list = []
            discount_price_list = []
            for i in text:
                s = BeautifulSoup(str(i), 'lxml')
                price = int(re.sub(r"[\s,₽]", "", s.find('span',
                                                         {'class': 'product-offer-price__amount'}).get_text(
                    strip=True)))
                bonus = s.find('span', {'class': 'bonus-amount'})
                bonus = int(bonus.get_text(strip=True).replace(' ', '')) if bonus else 0
                price_list.append(price)
                discount_price_list.append(price - (bonus * 0.7))
            try:
                if min(price_list) < min(discount_price_list)+(min(discount_price_list)*0.35):
                    del self.products[name]
            except ValueError:
                pass
        self.save()

    def run(self):
        self.create_table()
        for i in range(0, 210, 30):
            self.counter = i
            self.calculation()


if __name__ =='__main__':
    urls = ['smartfony-android']
    for url in urls:
        print(Parse(f'https://megamarket.ru/catalog/{url}/', url).parse())