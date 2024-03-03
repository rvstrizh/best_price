from parse import Parse, Parse_Price
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium import webdriver

class Main:
    def __init__(self, url):
        self.dict_page = {}
        self.url = url
        self.count = 0
        global wb
        wb = load_workbook('price.xlsx')



    def open(self, brand):
        if brand in wb.sheetnames:
            wb.remove(wb[brand])
        wb.create_sheet(brand)
        wb.save('./price.xlsx')

    def save_phone(self):
        for brand in self.dict_page:
            self.open(brand)
            sheet = wb[f'{brand}']
            counter = 0
            temporary_dict = dict(sorted(self.dict_page[brand].items(), key=lambda item: -item[1][1]))
            for name, url_sale in temporary_dict.items():
                counter += 1
                sheet[f'A{counter}'], sheet[f'B{counter}'], sheet[f'C{counter}'], sheet[f'D{counter}'], sheet[
                    f'E{counter}'] = \
                    name, url_sale[0], url_sale[1], url_sale[2], url_sale[3]
            wb.save('price.xlsx')

    def save(self):
        if self.url == 'smartfony-android':
            self.save_phone()
        else:
            sheet = wb[f'{self.url}']
            counter = 0
            self.dict_page = dict(sorted(self.dict_page.items(), key=lambda item: -item[1][1]))
            for name, url_sale in self.dict_page.items():
                counter += 1
                sheet[f'A{counter}'], sheet[f'B{counter}'], sheet[f'C{counter}'], sheet[f'D{counter}'], sheet[f'E{counter}'] = \
                    name, url_sale[0], url_sale[1], url_sale[2], url_sale[3]
            wb.save('price.xlsx')

    def paginator_page(self):
        url = f'https://megamarket.ru/catalog/{self.url}/'
        dict_ = Parse(url, self.url).parse()
        self.dict_page |= dict_
        count = 1
        while True:
        # for i in range(2):
            count += 1
            try:
                dict_ = Parse(f'{url}page-{count}', self.url).parse()
                if self.url == 'smartfony-android':
                    for brand in dict_:
                        if brand in self.dict_page:
                            self.dict_page[brand] |= dict_[brand]
                        else:
                            self.dict_page[brand] = dict_[brand]
                    print(count)
                else:
                    self.dict_page |= dict_
                print('test=', count)
                if count >= 50:
                    break
            except TypeError:
                break

    def start(self):
        self.paginator_page()
        self.open(self.url)
        self.save()
        # Parse_Price(self.url)


if __name__ =='__main__':
    urls = ['naushniki', 'umnye-chasy', 'portativnye-kolonki', 'televizory', 'smartfony-android', 'smartfony-android']
    # urls = ['kofemashiny', 'planshety', 'blendery', 'chajniki-elektricheskie', 'multirezki', 'kuhonnye-kombajny-i-mashiny', 'konstruktory-lego', 'igrovye-pristavki-playstation', 'igrovye-pristavki-xbox', 'portativnye-igrovye-konsoli', 'geympady', 'elektricheskie-zubnye-shetki', 'irrigatory']
    # urls = ['naushniki', 'smartfony-android']
    for url in urls:
        try:
            
            Main(url).start()
        except TimeoutException:
            Main(url).start()

    # Parse_Price(url).run()
